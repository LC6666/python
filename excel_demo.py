# -*coding:utf-8 -*-
__author__="豆豆嗯嗯"

from openpyxl import Workbook
from openpyxl import load_workbook

if __name__ == "__main__":

    print("写excel简单示例")

    #创建一个excel工作区
    wb = Workbook()

    #激活当前工作区
    ws = wb.active

    #往单元格A1写入数据，其他单元格写入类似
    ws['A1'] = "豆豆嗯嗯"

    #写下一行数据，列表元素对应每一个单元格
    ws.append([1,2,3])

    #写入时间类型到Excel,python会自动将类型换成excel的日期时间类型
    import datetime
    ws['A2'] = datetime.datetime.now()

    #保存为excel文件
    wb.save("豆豆.xlsx")

    print("读取已存在的excel文件")

    wb2 = load_workbook("豆豆.xlsx")

    #获取所有的sheet名，返回的是list类型
    sheets = wb2.get_sheet_names()
    print(type(sheets))

    #遍历sheets，并读取其单元格内容打印输入
    for sh in sheets:
        print("读取工作簿名称：",sh)

    ws = wb2.get_sheet_by_name(sheets[0])
    A1 = ws['A1'].value
    print("A1单元格的值：",A1)

    # 读取A2, B2, C2
    for index in ('A2', 'B2', 'C2'):
        print(index, "单元格的值： ", ws[index].value)

        # 读取空值的单元格, openpyxl对于空值的单元格，返回None
    F1 = ws['F1'].value
    print("F1单元格的值： ", F1)

    import urllib.request
    print ("爬取豆瓣网书籍数据写入excel示例")
    # 通过豆瓣网搜索API，搜索python关键词，采集书籍数据
    # 本示例只采集第一页的数据
    url = "https://api.douban.com/v2/book/search?q=python"
    response = urllib.request.urlopen(url)

    # 将bytes数据流解码成string
    ebook_str = response.read().decode()

    # 将string转换成dict
    ebook_dict = eval(ebook_str)

    # 构建一个Workbook对象
    wb = Workbook()  # 激活第一个sheet
    ws = wb.active  # 写入表头
    ws.append(["书名", "作者", "描述", "出版社", "价格"])

    # 写入书信息
    for book in ebook_dict["books"]:
        ws.append([book["title"],
                   ",".join(book["author"]),
                   book["summary"],
                   book["publisher"],
                   book["price"]])

        # 保存
    wb.save("ebook.xlsx")
    print ("数据保存完毕")
